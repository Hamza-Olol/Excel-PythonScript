from openpyxl.workbook import Workbook
from openpyxl import load_workbook

# Create a workbook object to load an existing excel file from the relative path

wb = load_workbook('people.xlsx')

# Load and active worksheet

ws = wb.active

# Select cells and print data from specified spreadsheet

name = ws["F3": "F10"]

for cell in name:
    for x in cell:
        print(x.value)